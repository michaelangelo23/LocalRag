import fitz
from typing import List
from langchain_text_splitters import RecursiveCharacterTextSplitter
import config 
import os
import traceback

from docx import Document as DocxDocument # To avoid naming conflict with fitz.Document
import openpyxl # For reading .xlsx files


def load_pdf_text(pdf_path: str) -> str:
    text = ""
    try:
        document = fitz.open(pdf_path)
        for page_num in range(document.page_count):
            page = document.load_page(page_num)
            # Directly extract text; no fallback to OCR
            page_text = page.get_text()
            text += page_text + "\n\n" # Add a separator between pages
        document.close()
    except Exception as e:
        print(f"Error loading PDF text from {os.path.basename(pdf_path)}: {e}")
        traceback.print_exc()
        return ""
    return text

def load_text_file_content(file_path: str) -> str:
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content
    except Exception as e:
        print(f"Error loading text file content from {os.path.basename(file_path)}: {e}")
        traceback.print_exc()
        return ""

def load_docx_text(docx_path: str) -> str:
    text = ""
    try:
        document = DocxDocument(docx_path)
        for paragraph in document.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        print(f"Error loading DOCX text from {os.path.basename(docx_path)}: {e}")
        traceback.print_exc()
        return ""

def load_xlsx_text(xlsx_path: str) -> str:
    text = ""
    try:
        workbook = openpyxl.load_workbook(xlsx_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"--- Sheet: {sheet_name} ---\n" # Header for each sheet
            for row in sheet.iter_rows():
                row_values = []
                for cell in row:
                    if cell.value is not None:
                        # Convert all cell values to string
                        row_values.append(str(cell.value).strip())
                if row_values:
                    text += "\t".join(row_values) + "\n" # Use tab for column separation
            text += "\n" # Add a newline between sheets
        return text
    except Exception as e:
        print(f"Error loading XLSX text from {os.path.basename(xlsx_path)}: {e}")
        traceback.print_exc()
        return ""

def split_text_into_chunks(text: str) -> List[str]:
    if not text.strip(): # Check for empty or whitespace-only text
        return []

    text_splitter = RecursiveCharacterTextSplitter(
        # Use values from config.py
        chunk_size=config.TEXT_CHUNK_SIZE,
        chunk_overlap=config.TEXT_CHUNK_OVERLAP,
        length_function=len,
        is_separator_regex=False, # Use standard separators
    )
    chunks = text_splitter.split_text(text)
    return chunks